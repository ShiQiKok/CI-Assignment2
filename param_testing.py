from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.impute import KNNImputer
from sklearn.preprocessing import OrdinalEncoder
from sklearn.neural_network import MLPClassifier
from sklearn.metrics import confusion_matrix, classification_report
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

def visualise(mlp):
    # get number of neurons in each layer
    n_neurons = [len(layer) for layer in mlp.coefs_]
    n_neurons.append(mlp.n_outputs_)

    # calculate the coordinates of each neuron on the graph
    y_range = [0, max(n_neurons)]
    x_range = [0, len(n_neurons)]
    loc_neurons = [[[l, (n+1)*(y_range[1]/(layer+1))] for n in range(layer)] for l,layer in enumerate(n_neurons)]
    x_neurons = [x for layer in loc_neurons for x,y in layer]
    y_neurons = [y for layer in loc_neurons for x,y in layer]

    # identify the range of weights
    weight_range = [min([layer.min() for layer in mlp.coefs_]), max([layer.max() for layer in mlp.coefs_])]

    # prepare the figure
    fig = plt.figure()
    ax = fig.add_subplot(1,1,1)
    # draw the neurons
    ax.scatter(x_neurons, y_neurons, s=100, zorder=5)
    # draw the connections with line width corresponds to the weight of the connection
    for l,layer in enumerate(mlp.coefs_):
        for i,neuron in enumerate(layer):
            for j,w in enumerate(neuron):
                ax.plot([loc_neurons[l][i][0], loc_neurons[l+1][j][0]], [loc_neurons[l][i][1], loc_neurons[l+1][j][1]], 'white', linewidth=((w-weight_range[0])/(weight_range[1]-weight_range[0])*5+0.2)*1.2)
                ax.plot([loc_neurons[l][i][0], loc_neurons[l+1][j][0]], [loc_neurons[l][i][1], loc_neurons[l+1][j][1]], 'grey', linewidth=(w-weight_range[0])/(weight_range[1]-weight_range[0])*5+0.2)


data = pd.read_csv('heart.csv')
categorical_input = ['Sex' , 'ChestPainType', 'RestingECG', 'ExerciseAngina', 'ST_Slope']
# print(data.describe().transpose())

## Encode the categorical data
oe = OrdinalEncoder()
oe.fit(data[categorical_input])
data[categorical_input] = oe.transform(data[categorical_input])
print(oe.categories_)

## Drop the data having RestingBP = 0
data.drop(data[data['RestingBP'] == 0].index, inplace=True)

## Impute value 0 for cholesterol
## replace value 0 to NaN for cholesterol
data['Cholesterol'].replace(0, np.nan, inplace=True)

## KNN Imputer will only impute NaN values
imputer = KNNImputer()
data = pd.DataFrame(imputer.fit_transform(data), columns=data.columns)


X = data.drop(columns='HeartDisease')
Y = data['HeartDisease']
X_train, X_test, y_train, y_test = train_test_split(X, Y, train_size=0.8, random_state=1)

# Scale the input
scaler = StandardScaler()
scaler.fit(X_train)
X_train = scaler.transform(X_train)
X_test = scaler.transform(X_test)

# ==========================Parameter Optimization ==========================
from openpyxl import load_workbook

def test_one_hidden_layer():    
    wb = load_workbook(filename="result.xlsx")
    sheet = wb['1layer']
    sheet['A1'] = 'Number of Nuerons'
    sheet['B1'] = 'Average Accuracy (10 iterations)'
    row = 2
    for n in range(1, 12):
        total = 0
        for iteration in range(10):
            mlp = MLPClassifier(hidden_layer_sizes=(n,), max_iter=1000)
            mlp.fit(X_train, y_train)
            predictions = mlp.predict(X_test)
            cr = classification_report(y_test, predictions,output_dict=True)
            accuracy = round(cr["accuracy"],2)
            total += accuracy
            print(f"layer: 1, neurons: {n}")
            print(accuracy)
        average = total / 10
        cell = sheet.cell(row, 1)
        cell.value = n
        cell = sheet.cell(row, 2)
        cell.value = average
        row += 1
    wb.save(filename="result.xlsx")
    
def test_two_hidden_layers():    
    wb = load_workbook(filename="result.xlsx")
    sheet = wb['2layers']
    sheet['A1'] = 'Number of Nuerons'
    sheet['B1'] = 'Average Accuracy (10 iterations)'
    row = 2
    for i in range(1, 12):
        for j in range(1, 12):
            total = 0
            for iteration in range(10):
                mlp = MLPClassifier(hidden_layer_sizes=(i, j,), max_iter=1000)
                mlp.fit(X_train, y_train)
                predictions = mlp.predict(X_test)
                cr = classification_report(y_test, predictions,output_dict=True)
                accuracy = round(cr["accuracy"],2)
                total += accuracy
                print(f"layer:2, neurons: ({i},{j})")
                print(accuracy)
            average = total / 10
            cell = sheet.cell(row, 1)
            cell.value = f"{i},{j}"
            cell = sheet.cell(row, 2)
            cell.value = average
            row += 1
    wb.save(filename="result.xlsx")

def test_activation(sizes=(4,)):
    wb = load_workbook(filename="result.xlsx")
    sheet = wb['activation']
    sheet['A1'] = f'Topology: {sizes}'
    sheet['A2'] = 'Activation'
    sheet['B2'] = 'Average Accuracy (10 iterations)'
    activation=["identity", "logistic", "tanh", "relu"]
    row = 3
    for a in activation:
        total = 0
        for iteration in range(10):
            mlp = MLPClassifier(hidden_layer_sizes=sizes, max_iter=1000)
            mlp.fit(X_train, y_train)
            predictions = mlp.predict(X_test)
            cr = classification_report(y_test, predictions,output_dict=True)
            accuracy = round(cr["accuracy"],2)
            total += accuracy
            print(f"activation: {a}, neurons: {sizes}")
            print(accuracy)
        average = total / 10
        cell = sheet.cell(row, 1)
        cell.value = a
        cell = sheet.cell(row, 2)
        cell.value = average
        row += 1
    wb.save(filename="result.xlsx")
    




